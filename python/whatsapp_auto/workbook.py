# Criar e adiciona arquivos em um excel

# importa a biblioteca para ler o excel
from openpyxl import Workbook

# varialvel que armazenha os excel
arquivo_excel = Workbook()

# ativa a planilha padrão
planilha1 = arquivo_excel.active

# colocando um nome para a planilha padrão
planilha1.title = "Mensagens"

# comandos para criar uma nova planilha
# planilha2 = arquivo_excel.create_sheet("nome")
# planilha2 = arquivo_excel.create_sheet("nome", 0)

# mostra as planilhas criadas e na ordem em que elas estão
print(arquivo_excel.sheetnames)

# -------  Adicionando dados à uma planilha --------#
planilha1['A1'] = "mensagem"
planilha1['B1'] = "numero"
planilha1['C1'] = "data"
planilha1['D1'] = "hora"
planilha1['E1'] = "arquivo"

planilha1['A2'] = "Olá"
planilha1['B2'] = "991180208"
planilha1['C2'] = "12/05"
planilha1['D2'] = "15:00"
planilha1['E2'] = "null"


